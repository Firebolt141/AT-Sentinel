"""
================================================================================
SAP UAT Test Management Automation
================================================================================
Version  : 2.1
Author   : Adarsh A. Shetty
Platform : Windows (local OneDrive sync + Outlook / Teams)

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------
Automates the daily overhead of managing ~500 SAP UAT test cycles across
domestic (Japan) and international (India) sales-order scenarios.

Problems this script solves:
  #1  Actual execution dates are never entered by executors in real time.
      → Reads condition script files and auto-populates actual dates when all
        steps are marked OK. No manual entry required.
  #2  False "Complete" status without script completion (regression risk).
      → Execution status is derived solely from script content, not manual input.
  #3  2+ hours/day in standups and 20+ individual chats to collect status.
      → Daily summary report is generated and distributed automatically at 8am.
  #4  No visibility into upcoming tests; prep reminders are missed.
      → Flags cycles starting within N days; alerts for overdue and stalled.
  #5  Scheduling conflicts (multiple cycles on same day/executor) are invisible.
      → Report surfaces workload concentration per executor per day.

--------------------------------------------------------------------------------
END-TO-END FLOW
--------------------------------------------------------------------------------
  1. DISCOVER FILES
     Scans the configured local folder (OneDrive-synced path) for:
       • One cycle list file  — filename matches regex: サイクル一覧
       • N condition scripts  — filename matches regex: 仕様書兼結果記述書.*AT1TC
     Only .xlsx files are considered. Excel lock files (~$*) are skipped.

  2. ANALYZE CONDITION SCRIPTS  (analyze_script)
     For each condition script file:
       a. Opens the sheet "テスト仕様書兼結果記述書"
       b. Skips header rows (rows 0–3) and excluded steps (col 8 = 'X')
       c. Stops processing at sentinel row where col 0 = 'e'
       d. A row is a valid step if it has a step number (col 1) OR result data
          — handles merged cells where step number only appears in first merged row
       e. Counts total active steps vs completed (result = OK) steps
       f. Collects first and last actual execution date from OK rows
          — retest date (col 32) takes priority over first-run date (col 27)
       g. Collects earliest planned execution date from any step (col 26)
       h. Detects NG results (regression indicator)
       i. Extracts executor names for backfill into cycle list

  3. UPDATE CYCLE LIST  (update_cycle_list)
     Opens テストサイクル一覧 sheet (read-write, formulas preserved).
     For each data row (from row 4 onward):
       a. Reconstructs cycle ID from area (col H) + seq_no (col C)
          The formula in col A cannot be evaluated by openpyxl, so the ID is
          rebuilt using the same logic: f"AT1TC_{area}{seq_no:03d}"
       b. Skips rows where deletion flag (col J / index 9) = 'X'
          NOTE: col I / index 8 is the REGRESSION flag — do not confuse the two
       c. Looks up the matching script result by cycle ID
       d. Updates these columns (guarded as noted):
            plan_start_latest (col 24) ← script's earliest planned date  [only if blank]
            actual_start      (col 26) ← script's first OK actual date   [only if blank]
            actual_end        (col 27) ← script's last OK actual date    [only if blank, all_ok only]
            exec_status       (col 33) ← derived from completion state   [always overwritten]
            total_steps       (col 34) ← count of active steps           [always overwritten]
            complete_steps    (col 35) ← count of OK steps               [always overwritten]
            executor          (col 32) ← from script's exec_pic column   [only if blank]
       e. Saves the workbook in place — OneDrive syncs it back to SharePoint

  4. DETECT REMINDERS  (get_reminders)
     Re-reads the (now updated) cycle list and flags:
       • upcoming — plan_start_latest within CONFIG["reminder_days_ahead"] days,
                    cycle not yet started
       • overdue  — plan_end_latest has passed, cycle not complete
       • stalled  — in-progress, plan_end within last 5 days, no actual_end set

  5. SEND NOTIFICATIONS  (notify)
     Dispatches to one or both channels based on CONFIG["notify_channels"]:

       "email"
         Sends via the locally installed Outlook app using win32com.
         No API keys or auth tokens — uses your already-logged-in session.
         Executor reminders go to individual executor email addresses.
         In test_mode, all executor mail is redirected to manager_email.

       "teams"
         Posts to a Teams channel via its built-in email address.
         Every Teams channel has a unique email (find it via channel → ••• →
         "Get email address"). Messages appear as email cards in the channel.
         No webhook, no Azure AD app registration, no IT approval required.
         Store the channel address in CONFIG["teams_channel_email"].

       Both channels can be active simultaneously.
       CONFIG["notify_channels"] = ["email", "teams"]  → both
       CONFIG["notify_channels"] = ["email"]            → email only
       CONFIG["notify_channels"] = ["teams"]            → Teams only

       UI NOTE: A future UI layer will expose channel selection per-run.
       The config values here serve as the default / fallback.

  6. REPORTS
     • Manager daily summary — HTML table: completed / started / overdue /
                               stalled / upcoming cycles
     • Executor reminders    — per-executor HTML message per reminder type
                               (upcoming / overdue / stalled)

--------------------------------------------------------------------------------
FILE NAMING CONVENTIONS
--------------------------------------------------------------------------------
  Cycle list (one per folder):
    Must contain:  サイクル一覧
    Example:       AT1_SAP内結合テストサイクル一覧_v1.0.xlsx

  Condition scripts (one per cycle):
    Must contain:  仕様書兼結果記述書  AND  AT1TC
    Example:       ⑧_TE586_SAP内結合テスト仕様書兼結果記述書_AT1TC_SDFI001.xlsx

  Cycle ID mapping (filename → cycle list):
    Filename uses  "SDFI001"  but cycle list stores  "SD/FI001"
    The slash is inserted automatically:
      if len(area_code) > 2 → area_code[:2] + "/" + area_code[2:]
    Examples:  SDFI → SD/FI  |  MM → MM  |  FI → FI

--------------------------------------------------------------------------------
COLUMN MAPPING — テストサイクル一覧  (0-indexed)
--------------------------------------------------------------------------------
  Index  0  (col A) : テストサイクルID       — FORMULA; reconstructed from H + C
  Index  1  (col B) : サイクル名称
  Index  2  (col C) : 領域内サイクル連番     — used in cycle ID reconstruction
  Index  7  (col H) : 担当領域               — area code (SD/FI, MM, FI, etc.)
  Index  8  (col I) : リグレッション対象フラグ — 'X' = regression target (NOT deletion)
  Index  9  (col J) : 削除フラグ             — 'X' = deleted; row is skipped entirely
  Index 24  (col Y) : 実行開始予定日(最新)   — plan_start_latest; backfilled from script
  Index 25  (col Z) : 実行完了予定日(最新)   — plan_end_latest; not auto-populated
  Index 26  (col AA): 実行開始実績日         — actual_start; set from first OK date
  Index 27  (col AB): 実行完了実績日         — actual_end; set from last OK date (all_ok)
  Index 28  (col AC): レビュー完了実績日     — review_end; set when all review dates filled [NEW]
  Index 29  (col AD): レビュー完了予定日     — review_plan_end; planned review end (optional) [NEW]
  Index 32  (col AG): 実行担当者             — executor; backfilled from script
  Index 33  (col AH): 実行ステータス         — exec_status; always synced by automation
  Index 34  (col AI): 総テストステップ数     — total_steps; always synced
  Index 35  (col AJ): 完了テストステップ数   — complete_steps; always synced

  Header rows: 3 (data starts at Excel row 4 / openpyxl min_row=4)

--------------------------------------------------------------------------------
COLUMN MAPPING — テスト仕様書兼結果記述書  (0-indexed)
--------------------------------------------------------------------------------
  Index  8  (col I) : テスト対象外           — 'X' = skip this step entirely
  Index 25  (col Z) : 初回実行 担当者        — executor name
  Index 26  (col AA): 初回実行 実行予定日    — planned_date; feeds earliest_planned
  Index 27  (col AB): 初回実行 実行日        — actual_date; feeds first/last_actual_date
  Index 28  (col AC): 初回実行 テスト結果    — OK / NG / -
  Index 30  (col AE): 再実行 担当者
  Index 31  (col AF): 再実行 実行予定日
  Index 32  (col AG): 再実行 実行日          — takes priority over first-run actual_date
  Index 33  (col AH): 再実行 テスト結果      — takes priority over first-run result
  Index 35  (col AJ): レビュー担当者         — reviewer name (placeholder; verify per stage) [NEW]
  Index 36  (col AK): レビュー実施日         — actual review date [NEW]
  Index 37  (col AL): レビュー結果           — review result (OK/NG/-) [NEW]

  Header rows: 4 (data starts at Excel row 5 / row_idx >= 4 in iter_rows)
  Sentinel:   Processing stops when col 0 = 'e'

--------------------------------------------------------------------------------
EXECUTION STATUS VALUES  (dropdown from the 'list' sheet)
--------------------------------------------------------------------------------
  00.未開始    — Not started   (no actual dates entered on any step)
  10.実行中    — In progress   (≥1 step has an actual date, OR any NG)
  20.レビュー中 — In Review    (all execution dates filled + all results OK; set by automation)
  SIR         — SIR           (≥1 step result = NG; regression found)
  30.完了      — Complete      (all steps have a review completion date)
  99.キャンセル — Cancelled    (excluded from all automation processing)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------
  Python   : 3.10+  (uses X | Y union type hints)
  Packages : pip install openpyxl pywin32
  OS       : Windows only (win32com requires the Windows COM interface)
  Outlook  : Must be installed locally and signed in
  OneDrive : Test file folder must be synced to a local path

--------------------------------------------------------------------------------
SCHEDULING  (Windows Task Scheduler — recommended setup)
--------------------------------------------------------------------------------
  Action   : Start a program
  Program  : C:\path\to\python.exe
  Arguments: "C:\path\to\this_script.py"
  Start in : C:\path\to\script_folder
  Trigger  : Daily at 08:00
  Settings : Run whether user is logged on or not
             Wake computer to run this task
             Run task as soon as possible after a scheduled start is missed

--------------------------------------------------------------------------------
KNOWN LIMITATIONS
--------------------------------------------------------------------------------
  • openpyxl cannot evaluate Excel formulas, so the cycle ID in col A (which
    is a formula) is rebuilt from area (col H) and seq_no (col C) instead.
    If the formula logic changes in the Excel file, update reconstruct_cycle_id.

  • Merged cells in condition scripts: if a step spans multiple rows via merge,
    openpyxl returns None for the cell values in non-anchor rows. The script
    handles this by accepting rows with either a step number OR result data,
    but very unusual merge layouts may miscount steps.

  • Teams channel email posts appear as email cards in the channel, not as
    native chat messages. For richer card formatting, replace the Teams sender
    with a Power Automate HTTP-triggered flow.

  • test_mode = True redirects ALL executor emails to manager_email.
    Set to False only after fully populating executor_emails in CONFIG.

  • plan_end_latest (col 25) is intentionally not auto-populated by this script.
    It is left for manual entry or a future enhancement.

--------------------------------------------------------------------------------
CHANGE LOG
--------------------------------------------------------------------------------
  v2.1  Added plan_start_latest backfill from script's earliest planned date
        Added Teams channel email notification (no webhook / API required)
        Added notify_channels config for dual email + Teams delivery
        Added notify() dispatcher; run() uses it instead of direct send calls
        Added this detailed header documentation

  v2.0  Full rewrite — fully local, no Azure / Graph API / Power Automate
        Cycle ID reconstructed from area + seq_no (formula workaround)
        Deletion flag (col 9) correctly distinguished from regression flag (col 8)
        Executor backfill, step count sync, retest-takes-priority logic

  v1.0  Initial prototype
================================================================================
"""

import json
import os
import re
import logging
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────
# CONFIG  ← edit these before first run
# ─────────────────────────────────────────────────────────────
CONFIG = {
    # Full path to the folder containing ALL test files
    "folder": r"C:\Users\adarsh.a.shetty\Downloads",

    # Filename patterns to identify files (regex)
    # Cycle list: contains "サイクル一覧" in the name
    "cycle_list_pattern": r"サイクル一覧",

    # Condition scripts: contains "仕様書兼結果記述書" AND "AT1TC" (to exclude unit test scripts)
    "condition_script_pattern": r"仕様書兼結果記述書.*AT1TC|AT1TC.*仕様書兼結果記述書",

    # How many days ahead to warn executors about upcoming cycles
    "reminder_days_ahead": 2,

    # Your email — ALL emails (manager report + executor reminders) go here during testing
    "manager_email": "adarsh.a.shetty@accenture.com",

    # TEST MODE: all executor reminder emails are redirected to manager_email above.
    # Set to False only when you are ready to send to real executors.
    "test_mode": True,

    # Executor name → email mapping
    # Not used while test_mode=True (all mail goes to manager_email instead)
    # Fill these in before going live.
    "executor_emails": {
        # "Suzuki":  "saeko.suzuki@accenture.com",
        # "鈴木":    "saeko.suzuki@accenture.com",
        # "Tanaka":  "tanaka.x@accenture.com",
        # "田中":    "tanaka.x@accenture.com",
    },

    # Log file location
    "log_file": r"C:\Users\adarsh.a.shetty\Downloads\sap_test_automation.log",

    # ── Notification channels ──────────────────────────────────────────────
    # List any combination of "email" and "teams".
    # "email"  → sends via local Outlook app (win32com, no API needed)
    # "teams"  → posts to a Teams channel via its email address (see below)
    # A future UI layer will let the user override this per run.
    "notify_channels": ["email", "teams"],

    # Teams channel email address.
    # How to find it: open the Teams channel → click ••• → "Get email address"
    # Leave blank ("") to disable Teams notifications even if "teams" is listed above.
    "teams_channel_email": "",

    # Dry-run mode: compute all changes but do NOT write the cycle list or send any
    # notifications. Useful for validating what the automation would do before going live.
    "dry_run": False,

    # Active test-stage profile key (see DEFAULT_PROFILES below).
    # Change this here for headless/scheduled runs, or use the Streamlit sidebar.
    "active_profile": "AT-SAP",
}

# Execution status dropdown values (from the 'list' sheet)
STATUS = {
    "not_started": "00.未開始",
    "in_progress": "10.実行中",
    "reviewing":   "20.レビュー中",
    "sir":         "SIR",
    "complete":    "30.完了",
    "cancelled":   "99.キャンセル",
}

# ─────────────────────────────────────────────────────────────
# PROFILES — per-test-stage / per-project configuration
# ─────────────────────────────────────────────────────────────
# Each profile captures everything that differs between test stages
# (AT-SAP, BLT, AT-IF, SIT) or between project customisations:
#   • filename patterns          • Excel sheet names
#   • header row counts          • all column indices (0-based)
#   • cycle-ID prefix & format
#
# AT-SAP is the fully-verified reference profile.
# BLT / AT-IF / SIT ship as templates that mirror AT-SAP column
# indices — verify against each stage's actual Excel files and
# save adjustments via the Streamlit "Edit Profile" panel.
# Saved changes are written to profiles.json next to this script
# and override the built-in defaults on every future run.
# ─────────────────────────────────────────────────────────────

PROFILES_FILE: Path = Path(__file__).parent / "profiles.json"

DEFAULT_PROFILES: dict[str, dict] = {
    "AT-SAP": {
        "display_name":             "AT-SAP (SAP内結合テスト)",
        "cycle_list_pattern":       r"サイクル一覧",
        "condition_script_pattern": r"仕様書兼結果記述書.*AT1TC|AT1TC.*仕様書兼結果記述書",
        "cycle_sheet_name":         "テストサイクル一覧",
        "script_sheet_name":        "テスト仕様書兼結果記述書",
        "cycle_header_rows":        3,
        "script_header_rows":       4,
        "cycle_id_prefix":          "AT1TC",
        "cycle_id_area_slash":      True,   # SDFI → SD/FI
        "configured":               True,   # set False on template profiles
        "cycle_cols": {
            "cycle_id":          0,
            "seq_no":            2,
            "cycle_name":        1,
            "area":              7,
            "regression_flag":   8,
            "deletion_flag":     9,
            "plan_start_latest": 24,
            "plan_end_latest":   25,
            "actual_start":      26,
            "actual_end":        27,
            "executor":          32,
            "exec_status":       33,
            "total_steps":       34,
            "complete_steps":    35,
            "review_end":        28,   # col AC: レビュー完了実績日 — review completion date
            "review_plan_end":   29,   # col AD: レビュー完了予定日 — planned review end (optional)
        },
        "script_cols": {
            "excluded":       8,
            "exec_pic":       25,
            "planned_date":   26,
            "actual_date":    27,
            "result":         28,
            "retest_pic":     30,
            "retest_planned": 31,
            "retest_date":    32,
            "retest_result":  33,
            "review_pic":     35,   # col AJ: レビュー担当者 — reviewer name (placeholder; verify per stage)
            "review_date":    36,   # col AK: レビュー実施日 — actual review date
            "review_result":  37,   # col AL: レビュー結果   — review result (OK/NG/-)
        },
    },
}

# Build BLT / AT-IF / SIT as template copies of AT-SAP.
# Column indices default to AT-SAP values — update via the UI.
for _stage, _prefix, _display, _pat in [
    ("BLT",   "BLT",  "BLT (業務連携テスト)",         r"仕様書兼結果記述書.*BLT|BLT.*仕様書兼結果記述書"),
    ("AT-IF", "ATIF", "AT-IF (インターフェーステスト)", r"仕様書兼結果記述書.*ATIF|ATIF.*仕様書兼結果記述書"),
    ("SIT",   "SIT",  "SIT (システム統合テスト)",       r"仕様書兼結果記述書.*SIT|SIT.*仕様書兼結果記述書"),
]:
    DEFAULT_PROFILES[_stage] = {
        **DEFAULT_PROFILES["AT-SAP"],
        "display_name":             f"{_display}  ⚠ verify column indices",
        "condition_script_pattern": _pat,
        "cycle_id_prefix":          _prefix,
        "configured":               False,
        "cycle_cols":               dict(DEFAULT_PROFILES["AT-SAP"]["cycle_cols"]),
        "script_cols":              dict(DEFAULT_PROFILES["AT-SAP"]["script_cols"]),
    }

# ─────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────
def setup_logging():
    # Guard against duplicate handlers when called multiple times (e.g. Streamlit reruns)
    root = logging.getLogger()
    if root.handlers:
        return
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(CONFIG["log_file"], encoding="utf-8"),
        ]
    )

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# PROFILE MANAGEMENT
# ─────────────────────────────────────────────────────────────
def load_profiles() -> dict[str, dict]:
    """
    Return the merged profile dict: built-in DEFAULT_PROFILES overridden by
    any user edits stored in profiles.json.

    Merging rules:
      • A profile that exists in profiles.json completely replaces the built-in
        for top-level scalar keys (display_name, patterns, sheet names, …).
      • cycle_cols / script_cols are merged at the field level so a partial
        override (e.g. only changing two column indices) leaves the rest intact.
      • A brand-new profile not present in DEFAULT_PROFILES is seeded from
        AT-SAP so it always has every required key.
    """
    if not PROFILES_FILE.exists():
        return {k: dict(v) for k, v in DEFAULT_PROFILES.items()}

    try:
        with open(PROFILES_FILE, "r", encoding="utf-8") as f:
            user_profiles: dict = json.load(f)
    except Exception as e:
        log.warning(f"Could not read {PROFILES_FILE.name}: {e} — using built-in defaults")
        return {k: dict(v) for k, v in DEFAULT_PROFILES.items()}

    merged: dict[str, dict] = {k: dict(v) for k, v in DEFAULT_PROFILES.items()}
    for name, user_p in user_profiles.items():
        base = dict(merged.get(name, DEFAULT_PROFILES["AT-SAP"]))
        for key, val in user_p.items():
            if key in ("cycle_cols", "script_cols") and isinstance(val, dict):
                base[key] = {**base.get(key, {}), **val}
            else:
                base[key] = val
        merged[name] = base
    return merged


def save_profiles(profiles: dict[str, dict]) -> None:
    """Persist the complete profiles dict to profiles.json."""
    try:
        with open(PROFILES_FILE, "w", encoding="utf-8") as f:
            json.dump(profiles, f, ensure_ascii=False, indent=2)
        log.info(f"Profiles saved → {PROFILES_FILE}")
    except Exception as e:
        log.error(f"Could not save profiles: {e}")


def get_active_profile() -> dict:
    """Return the profile dict for CONFIG['active_profile'], falling back to AT-SAP."""
    profiles = load_profiles()
    name = CONFIG.get("active_profile", "AT-SAP")
    if name not in profiles:
        log.warning(f"Profile '{name}' not found — falling back to AT-SAP")
        name = "AT-SAP"
    return profiles[name]


# ─────────────────────────────────────────────────────────────
# FILE DISCOVERY
# ─────────────────────────────────────────────────────────────
def find_files(folder: str, *, profile: dict | None = None) -> tuple[str | None, list[str]]:
    """
    Scan folder for:
    - One cycle list file  (matches profile's cycle_list_pattern)
    - N condition scripts  (match profile's condition_script_pattern)

    Returns (cycle_list_path, [script_paths])
    """
    p = profile or get_active_profile()
    cycle_list = None
    scripts = []
    root = Path(folder)

    for f in root.rglob("*.xlsx"):
        if f.name.startswith("~$"):  # skip Excel temp/lock files
            continue

        # Skip files inside any subfolder whose name contains "old" or "旧" (case-insensitive)
        relative_parts = f.relative_to(root).parts[:-1]  # intermediate folder names only
        if any("old" in part.lower() or "旧" in part for part in relative_parts):
            log.debug(f"Skipping (old/旧 folder): {f}")
            continue

        name = f.name
        if re.search(p["cycle_list_pattern"], name):
            cycle_list = str(f)
            log.info(f"Cycle list: {name}")
        elif re.search(p["condition_script_pattern"], name):
            scripts.append(str(f))
            log.info(f"Condition script: {name}")

    log.info(f"Found {len(scripts)} condition scripts")
    return cycle_list, scripts


def reconstruct_cycle_id(area: str, seq_no, *, profile: dict | None = None) -> str | None:
    """
    Reconstruct the cycle ID from area + seq_no — mirrors the Excel formula
      ="<prefix>_"&<area_col>&TEXT(<seq_col>,"000")
    e.g. prefix="AT1TC", area="SD/FI", seq_no=1 → "AT1TC_SD/FI001"

    The prefix and area-slash behaviour come from the active profile.
    """
    p = profile or get_active_profile()
    if not area or seq_no is None:
        return None
    try:
        area_str = str(area).strip()
        if p.get("cycle_id_area_slash", True) and len(area_str) > 2:
            area_str = area_str[:2] + "/" + area_str[2:]
        return f"{p['cycle_id_prefix']}_{area_str}{int(seq_no):03d}"
    except (ValueError, TypeError):
        return None


def extract_cycle_id_from_filename(filename: str, *, profile: dict | None = None) -> str | None:
    """
    Extract the cycle ID from a condition script filename using the profile's
    cycle_id_prefix.

    e.g. profile prefix="AT1TC", filename contains "AT1TC_SDFI001"
         → "AT1TC_SD/FI001"  (area slash inserted when cycle_id_area_slash=True)

    The regex is built from the profile prefix so BLT, ATIF, SIT etc. all work
    without code changes.
    """
    p = profile or get_active_profile()
    prefix_pat = re.escape(p["cycle_id_prefix"])
    match = re.search(rf'({prefix_pat})_([A-Z]+)(\d+)', filename)
    if not match:
        return None

    prefix = match.group(1)
    area   = match.group(2)
    num    = match.group(3)

    if p.get("cycle_id_area_slash", True) and len(area) > 2:
        area = area[:2] + "/" + area[2:]

    return f"{prefix}_{area}{num}"


# ─────────────────────────────────────────────────────────────
# CONDITION SCRIPT ANALYSIS
# ─────────────────────────────────────────────────────────────
def analyze_script(filepath: str, *, profile: dict | None = None) -> dict | None:
    """
    Read a condition script and return completion status.
    Sheet names, column indices, and header-row counts come from the profile.

    Returns dict with:
      cycle_id          - extracted from filename
      total_steps       - active (non-excluded) steps
      completed_steps   - steps with result = OK
      all_ok            - True if all active steps are OK and no NG
      has_ng            - True if any step has NG
      first_actual_date - earliest actual execution date
      last_actual_date  - latest actual execution date
      executor          - name(s) from 実行担当者 column
      earliest_planned  - earliest planned execution date (for reminders)
    """
    p = profile or get_active_profile()
    cycle_id = extract_cycle_id_from_filename(Path(filepath).name, profile=p)
    if not cycle_id:
        log.warning(f"Could not extract cycle ID from: {Path(filepath).name}")
        return None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"Could not open {Path(filepath).name}: {e}")
        return None

    sheet_name = p["script_sheet_name"]
    if sheet_name not in wb.sheetnames:
        log.warning(f"Sheet '{sheet_name}' not found in {Path(filepath).name}")
        return None

    ws = wb[sheet_name]
    c = p["script_cols"]
    hdr = p["script_header_rows"]

    total = 0
    completed = 0
    has_ng = False
    actual_dates = []
    ok_steps_missing_date = 0  # OK steps with no actual date (for all_dates_filled)
    planned_dates = []
    executors = set()
    last_step_no = None  # carries the step number forward across merged-cell rows
    any_actual_start = False  # any step with actual_date filled (triggers In Progress)
    review_dates: list = []
    steps_no_review = 0        # active steps missing a review date (for all_review_complete)

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < hdr:
            continue

        # Stop at sentinel row
        if row and row[0] == "e":
            break

        # Skip empty rows
        if not row or all(v is None for v in row):
            continue

        # Skip excluded steps (col 8 = テスト対象外)
        if len(row) > c["excluded"] and row[c["excluded"]] == "X":
            continue

        # Carry the step number forward across merged-cell rows.
        # In a merged step group only the anchor row has a value in col 1;
        # subsequent rows return None.  Updating last_step_no here ensures
        # every row in the group is treated as a valid step, even when untouched
        # (no result data yet).
        step_no = row[1] if len(row) > 1 else None
        if step_no is not None:
            last_step_no = step_no
        has_step_no = last_step_no is not None
        has_result_data = (
            len(row) > c["result"] and row[c["result"]] in ("OK", "NG", "-")
        ) or (
            len(row) > c["retest_result"] and row[c["retest_result"]] in ("OK", "NG", "-")
        ) or (
            len(row) > c["actual_date"] and row[c["actual_date"]] is not None
        ) or (
            len(row) > c["planned_date"] and row[c["planned_date"]] is not None
        )

        if not has_step_no and not has_result_data:
            continue

        total += 1

        # Collect executor name
        pic = row[c["exec_pic"]] if len(row) > c["exec_pic"] else None
        if pic and str(pic).strip() not in ("", "-"):
            executors.add(str(pic).strip())

        # Collect planned date
        planned = row[c["planned_date"]] if len(row) > c["planned_date"] else None
        if planned and isinstance(planned, (datetime, date)):
            d = planned.date() if isinstance(planned, datetime) else planned
            planned_dates.append(d)

        # Determine result — retest takes priority over first run
        retest_result = row[c["retest_result"]] if len(row) > c["retest_result"] else None
        first_result  = row[c["result"]]         if len(row) > c["result"]         else None
        result = (retest_result
                  if isinstance(retest_result, str) and retest_result.strip() not in ("", "-")
                  else first_result)

        retest_date = row[c["retest_date"]] if len(row) > c["retest_date"] else None
        first_date  = row[c["actual_date"]] if len(row) > c["actual_date"] else None
        actual = retest_date if retest_date else first_date

        # Flag if any step has been started (actual date filled, regardless of result)
        if actual and isinstance(actual, (datetime, date)):
            any_actual_start = True

        if result == "OK":
            completed += 1
            if actual and isinstance(actual, (datetime, date)):
                d = actual.date() if isinstance(actual, datetime) else actual
                actual_dates.append(d)
            else:
                ok_steps_missing_date += 1
        elif result == "NG":
            has_ng = True

        # Review date — collect for all_review_complete and last_review_date
        rev_date_col = c.get("review_date")
        raw_rev = row[rev_date_col] if (rev_date_col is not None and len(row) > rev_date_col) else None
        if isinstance(raw_rev, datetime):
            review_dates.append(raw_rev.date())
        elif isinstance(raw_rev, date):
            review_dates.append(raw_rev)
        else:
            steps_no_review += 1

    wb.close()

    return {
        "cycle_id":            cycle_id,
        "total_steps":         total,
        "completed_steps":     completed,
        "all_ok":              total > 0 and completed == total and not has_ng,
        "has_ng":              has_ng,
        "first_actual_date":   min(actual_dates)   if actual_dates   else None,
        "last_actual_date":    max(actual_dates)   if actual_dates   else None,
        "all_dates_filled":    completed > 0 and ok_steps_missing_date == 0,
        "executor":            ", ".join(sorted(executors)) if executors else None,
        "earliest_planned":    min(planned_dates)  if planned_dates  else None,
        # New fields
        "has_any_actual_start": any_actual_start,
        "all_review_complete":  total > 0 and steps_no_review == 0,
        "last_review_date":     max(review_dates)  if review_dates   else None,
    }


# ─────────────────────────────────────────────────────────────
# CYCLE LIST UPDATE
# ─────────────────────────────────────────────────────────────
def update_cycle_list(cycle_list_path: str, script_results: dict[str, dict], *,
                      profile: dict | None = None) -> tuple[list[dict], list[dict]]:
    """
    Open the Test Cycle List, update status/dates/counts from script results,
    save in place (unless dry_run is set in CONFIG).

    Returns (changes, missing_scripts):
      changes         — rows that were updated
      missing_scripts — active, non-cancelled cycles that have no matching script file

    IMPORTANT: We use load_workbook without data_only so formulas are preserved.
    We ONLY write to columns that contain plain values (not formulas).
    Formula columns (flags, report tabs) are left untouched.
    """
    p = profile or get_active_profile()
    wb = openpyxl.load_workbook(cycle_list_path)  # NOT read_only, NOT data_only
    ws = wb[p["cycle_sheet_name"]]
    c = p["cycle_cols"]
    hdr = p["cycle_header_rows"]
    today = date.today()
    changes: list[dict] = []
    missing_scripts: list[dict] = []
    seen_ids: set[str] = set()

    for row in ws.iter_rows(min_row=hdr + 1):
        # Reconstruct cycle ID from area + seq_no (col 0 is a formula openpyxl can't evaluate)
        area_val   = row[c["area"]].value
        seq_val    = row[c["seq_no"]].value
        cycle_id   = reconstruct_cycle_id(area_val, seq_val, profile=p)

        if not cycle_id:
            continue

        # Skip header remnants and sentinel rows
        raw_col0 = str(row[c["cycle_id"]].value or "").strip()
        if raw_col0 in ("e", "テストサイクルID", "Test cycle ID"):
            continue

        # Skip deleted cycles (col 9 = deletion flag, NOT col 8 which is regression flag)
        if row[c["deletion_flag"]].value == "X":
            continue

        # Skip cancelled cycles — automation must not overwrite their status
        current_status = row[c["exec_status"]].value or STATUS["not_started"]
        if current_status == STATUS["cancelled"]:
            continue

        # Warn and skip if two rows reconstruct to the same cycle ID
        if cycle_id in seen_ids:
            log.warning(f"Duplicate cycle ID '{cycle_id}' in cycle list — skipping row")
            continue
        seen_ids.add(cycle_id)

        result = script_results.get(cycle_id)
        if not result:
            # Active cycle with no matching condition script — surface for review
            missing_scripts.append({
                "cycle_id":   cycle_id,
                "cycle_name": row[c["cycle_name"]].value,
                "area":       row[c["area"]].value,
                "exec_status": current_status,
                "executor":   row[c["executor"]].value,
            })
            continue

        # Determine new status
        # Priority: complete > in_review > sir > in_progress > not_started
        # Backwards-compatible: fall back to completed_steps > 0 if result dict is old-style
        has_any_actual = result.get("has_any_actual_start", result["completed_steps"] > 0)
        all_review_done = result.get("all_review_complete", False)

        if all_review_done:
            new_status = STATUS["complete"]
        elif result["all_ok"] and result["all_dates_filled"]:
            new_status = STATUS["reviewing"]   # 20.レビュー中 — all execution done, pending review
        elif result["has_ng"]:
            new_status = STATUS["sir"]
        elif has_any_actual:
            new_status = STATUS["in_progress"]
        else:
            new_status = STATUS["not_started"]

        row_changed = False

        # planned start date — backfill from script if blank in ichiran
        if result["earliest_planned"] and not row[c["plan_start_latest"]].value:
            row[c["plan_start_latest"]].value = result["earliest_planned"]
            row_changed = True

        # actual start date — only set if blank
        if result["first_actual_date"] and not row[c["actual_start"]].value:
            row[c["actual_start"]].value = result["first_actual_date"]
            row_changed = True

        # actual end date — only set when all OK, all actual dates present, and currently blank
        if result["all_ok"] and result["all_dates_filled"] and result["last_actual_date"] and not row[c["actual_end"]].value:
            row[c["actual_end"]].value = result["last_actual_date"]
            row_changed = True

        # review end date — set when all steps have review dates and field is blank
        if ("review_end" in c
                and result.get("all_review_complete")
                and result.get("last_review_date")
                and not row[c["review_end"]].value):
            row[c["review_end"]].value = result["last_review_date"]
            row_changed = True

        # execution status
        if new_status != current_status:
            row[c["exec_status"]].value = new_status
            row_changed = True

        # step counts — always sync; flag row_changed when the numbers actually differ
        if (row[c["total_steps"]].value    != result["total_steps"] or
                row[c["complete_steps"]].value != result["completed_steps"]):
            row_changed = True
        row[c["total_steps"]].value    = result["total_steps"]
        row[c["complete_steps"]].value = result["completed_steps"]

        # executor — backfill if blank
        if result["executor"] and not row[c["executor"]].value:
            row[c["executor"]].value = result["executor"]
            row_changed = True

        if row_changed:
            plan_end = row[c["plan_end_latest"]].value
            if isinstance(plan_end, datetime):
                plan_end = plan_end.date()

            changes.append({
                "cycle_id":        cycle_id,
                "cycle_name":      row[c["cycle_name"]].value,
                "area":            row[c["area"]].value,
                "old_status":      current_status,
                "new_status":      new_status,
                "completed_steps": result["completed_steps"],
                "total_steps":     result["total_steps"],
                "executor":        result["executor"],
                "plan_end":        plan_end,
                "all_ok":          result["all_ok"],
                "has_ng":          result["has_ng"],
            })

    if CONFIG.get("dry_run"):
        log.info(f"DRY RUN — cycle list NOT saved ({len(changes)} changes computed, "
                 f"{len(missing_scripts)} missing scripts)")
    else:
        try:
            wb.save(cycle_list_path)
        except PermissionError:
            wb.close()
            raise PermissionError(
                f"Could not save cycle list — close the file in Excel first: {cycle_list_path}"
            )
        except Exception as e:
            wb.close()
            raise RuntimeError(f"Could not save cycle list: {e}") from e
    wb.close()
    log.info(f"Cycle list updated — {len(changes)} rows changed, "
             f"{len(missing_scripts)} cycles missing scripts")
    return changes, missing_scripts


# ─────────────────────────────────────────────────────────────
# REMINDER DETECTION
# ─────────────────────────────────────────────────────────────
def get_reminders(cycle_list_path: str, today: date, *, profile: dict | None = None,
                  time_slot: str = "morning") -> list[dict]:
    """
    Scan cycle list for cycles needing attention based on the notification time slot.

    time_slot values and the reminder types each produces:
      "morning" / "midday"  (9:00 and 12:00 runs)
        • "overdue"         — plan_end is today or in the past, cycle not complete
        • "starting_today"  — plan_start is today, cycle not yet started

      "evening"             (17:00 run — all of the above plus)
        • "starting_tomorrow" — plan_start is tomorrow, cycle not yet started
        • "due_tomorrow"      — plan_end is tomorrow, cycle not complete
    """
    p = profile or get_active_profile()
    wb = openpyxl.load_workbook(cycle_list_path, read_only=True, data_only=True)
    ws = wb[p["cycle_sheet_name"]]
    c = p["cycle_cols"]
    hdr = p["cycle_header_rows"]
    tomorrow = today + timedelta(days=1)
    reminders = []

    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row:
            continue

        area_val = row[c["area"]]
        seq_val  = row[c["seq_no"]]
        cycle_id = reconstruct_cycle_id(area_val, seq_val, profile=p)
        if not cycle_id:
            continue

        raw_col0 = str(row[c["cycle_id"]] or "").strip()
        if raw_col0 in ("e", "テストサイクルID", "Test cycle ID"):
            continue

        if row[c["deletion_flag"]] == "X":
            continue

        status     = row[c["exec_status"]] or STATUS["not_started"]
        plan_start = row[c["plan_start_latest"]]
        plan_end   = row[c["plan_end_latest"]]
        executor   = row[c["executor"]]

        if isinstance(plan_start, datetime): plan_start = plan_start.date()
        if isinstance(plan_end,   datetime): plan_end   = plan_end.date()

        if status in (STATUS["complete"], STATUS["cancelled"]):
            continue

        reminder_type = None

        # Always included (morning / midday / evening)
        if plan_end and plan_end <= today:
            # Plan end is today or overdue from previous days
            reminder_type = "overdue"
        elif plan_start and plan_start == today and status == STATUS["not_started"]:
            reminder_type = "starting_today"

        # Evening-only additions
        elif time_slot == "evening":
            if plan_start and plan_start == tomorrow and status == STATUS["not_started"]:
                reminder_type = "starting_tomorrow"
            elif plan_end and plan_end == tomorrow:
                reminder_type = "due_tomorrow"

        if reminder_type:
            reminders.append({
                "cycle_id":   cycle_id,
                "cycle_name": row[c["cycle_name"]],
                "executor":   executor,
                "plan_start": plan_start,
                "plan_end":   plan_end,
                "status":     status,
                "type":       reminder_type,
            })

    wb.close()
    return reminders


# ─────────────────────────────────────────────────────────────
# REPORT DATA HELPERS
# ─────────────────────────────────────────────────────────────
def _get_date_from_row(row: tuple, col_key: str, col_map: dict) -> date | None:
    """Safely extract a date value from a values_only row tuple by column key."""
    idx = col_map.get(col_key)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def read_cycle_data(cycle_list_path: str, *, profile: dict | None = None) -> list[dict]:
    """
    Read all active (non-deleted, non-cancelled) cycles from the cycle list.
    Returns a list of dicts with all fields needed for reporting.
    """
    p = profile or get_active_profile()
    wb = openpyxl.load_workbook(cycle_list_path, read_only=True, data_only=True)
    ws = wb[p["cycle_sheet_name"]]
    c = p["cycle_cols"]
    hdr = p["cycle_header_rows"]
    cycles: list[dict] = []

    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row:
            continue

        area_val = row[c["area"]] if c["area"] < len(row) else None
        seq_val  = row[c["seq_no"]] if c["seq_no"] < len(row) else None
        cycle_id = reconstruct_cycle_id(area_val, seq_val, profile=p)
        if not cycle_id:
            continue

        raw_col0 = str(row[c["cycle_id"]] or "").strip()
        if raw_col0 in ("e", "テストサイクルID", "Test cycle ID"):
            continue

        if row[c["deletion_flag"]] == "X":
            continue

        status = row[c["exec_status"]] or STATUS["not_started"]
        if status == STATUS["cancelled"]:
            continue

        executor_raw = row[c["executor"]] if c["executor"] < len(row) else None
        cycles.append({
            "cycle_id":       cycle_id,
            "cycle_name":     row[c["cycle_name"]] if c["cycle_name"] < len(row) else None,
            "area":           str(area_val).strip() if area_val else "Unknown",
            "executor":       str(executor_raw).strip() if executor_raw else None,
            "exec_status":    status,
            "plan_start":     _get_date_from_row(row, "plan_start_latest", c),
            "plan_end":       _get_date_from_row(row, "plan_end_latest",   c),
            "actual_start":   _get_date_from_row(row, "actual_start",      c),
            "actual_end":     _get_date_from_row(row, "actual_end",        c),
            "review_end":     _get_date_from_row(row, "review_end",        c),
            "review_plan_end":_get_date_from_row(row, "review_plan_end",   c),
            "total_steps":    row[c["total_steps"]]    if c["total_steps"]    < len(row) else 0,
            "complete_steps": row[c["complete_steps"]] if c["complete_steps"] < len(row) else 0,
        })

    wb.close()
    return cycles


def build_daywise_report(cycle_list_path: str, *, profile: dict | None = None) -> dict:
    """
    Build cumulative day-by-day execution progress data for charting.

    Returns:
        {
          "total":  int,
          "dates":  list[str]  (MM/DD formatted, sorted),
          "series": {series_label: list[int]}  (cumulative count per date)
        }
    Series labels: Start Plan, Start Actual, Comp Plan(Exe), Comp Actual(Exe),
                   Comp Plan(Review) [if review_plan_end present], Comp Actual(Review)
    """
    cycles = read_cycle_data(cycle_list_path, profile=profile)
    total = len(cycles)
    if not total:
        return {"total": 0, "dates": [], "series": {}}

    # Collect all meaningful dates to define the x-axis range
    all_dates: set[date] = set()
    for cyc in cycles:
        for d in (cyc["plan_start"], cyc["plan_end"], cyc["actual_start"],
                  cyc["actual_end"], cyc["review_end"], cyc["review_plan_end"]):
            if d:
                all_dates.add(d)

    if not all_dates:
        return {"total": total, "dates": [], "series": {}}

    sorted_dates = sorted(all_dates)
    has_review_plan = any(cyc["review_plan_end"] for cyc in cycles)

    series: dict[str, list[int]] = {
        "Start Plan":          [],
        "Start Actual":        [],
        "Comp Plan(Exe)":      [],
        "Comp Actual(Exe)":    [],
        "Comp Actual(Review)": [],
    }
    if has_review_plan:
        series["Comp Plan(Review)"] = []

    for d in sorted_dates:
        series["Start Plan"].append(
            sum(1 for cyc in cycles if cyc["plan_start"] and cyc["plan_start"] <= d))
        series["Start Actual"].append(
            sum(1 for cyc in cycles if cyc["actual_start"] and cyc["actual_start"] <= d))
        series["Comp Plan(Exe)"].append(
            sum(1 for cyc in cycles if cyc["plan_end"] and cyc["plan_end"] <= d))
        series["Comp Actual(Exe)"].append(
            sum(1 for cyc in cycles if cyc["actual_end"] and cyc["actual_end"] <= d))
        series["Comp Actual(Review)"].append(
            sum(1 for cyc in cycles if cyc["review_end"] and cyc["review_end"] <= d))
        if has_review_plan:
            series["Comp Plan(Review)"].append(
                sum(1 for cyc in cycles if cyc["review_plan_end"] and cyc["review_plan_end"] <= d))

    return {
        "total":  total,
        "dates":  [d.strftime("%m/%d") for d in sorted_dates],
        "series": series,
    }


def _workstream_summary(cycles: list[dict], group_key: str) -> list[dict]:
    """
    Shared logic for streamwise / executorwise reports.
    Groups cycles by `group_key` and computes execution + review metrics.
    """
    today = date.today()
    groups: dict[str, list[dict]] = {}
    for cyc in cycles:
        key = cyc.get(group_key) or "Unknown"
        groups.setdefault(key, []).append(cyc)

    rows: list[dict] = []
    totals = {"total": 0, "exec_plan": 0, "exec_actual": 0,
              "exec_ahead": 0, "exec_delay": 0,
              "review_plan": 0, "review_actual": 0,
              "review_ahead": 0, "review_delay": 0}

    for group_name in sorted(groups):
        grp = groups[group_name]
        total = len(grp)
        # Execution — "plan" = planned to complete by today, not all that have a plan_end
        exec_plan   = sum(1 for cyc in grp if cyc["plan_end"] and cyc["plan_end"] <= today)
        exec_actual = sum(1 for cyc in grp if cyc["actual_end"])
        exec_ahead  = sum(1 for cyc in grp
                          if cyc["actual_end"] and cyc["plan_end"]
                          and cyc["actual_end"] <= cyc["plan_end"])
        exec_delay  = sum(1 for cyc in grp
                          if (not cyc["actual_end"] and cyc["plan_end"] and cyc["plan_end"] < today)
                          or (cyc["actual_end"] and cyc["plan_end"]
                              and cyc["actual_end"] > cyc["plan_end"]))
        # Review
        review_plan  = sum(1 for cyc in grp
                           if cyc["review_plan_end"] and cyc["review_plan_end"] <= today)
        review_actual = sum(1 for cyc in grp if cyc["review_end"])
        review_ahead  = sum(1 for cyc in grp
                            if cyc["review_end"] and cyc["review_plan_end"]
                            and cyc["review_end"] <= cyc["review_plan_end"])
        review_delay  = sum(1 for cyc in grp
                            if (not cyc["review_end"] and cyc["review_plan_end"]
                                and cyc["review_plan_end"] < today)
                            or (cyc["review_end"] and cyc["review_plan_end"]
                                and cyc["review_end"] > cyc["review_plan_end"]))

        row = {
            group_key:       group_name,
            "total":         total,
            "exec_plan":     exec_plan,
            "exec_actual":   exec_actual,
            "exec_ahead":    exec_ahead,
            "exec_delay":    exec_delay,
            "review_plan":   review_plan,
            "review_actual": review_actual,
            "review_ahead":  review_ahead,
            "review_delay":  review_delay,
        }
        rows.append(row)
        for k in ("total", "exec_plan", "exec_actual", "exec_ahead", "exec_delay",
                  "review_plan", "review_actual", "review_ahead", "review_delay"):
            totals[k] += row[k]

    totals[group_key] = "Total"
    rows.append(totals)
    return rows


def build_streamwise_report(cycle_list_path: str, *, profile: dict | None = None) -> list[dict]:
    """
    Workstream (area) breakdown: total cycles vs execution/review plan & actual counts.
    Each row: area, total, exec_plan, exec_actual, exec_ahead, exec_delay,
              review_actual, review_ahead, review_delay.
    Last row is the "Total" summary.
    """
    cycles = read_cycle_data(cycle_list_path, profile=profile)
    return _workstream_summary(cycles, "area")


def build_executorwise_report(cycle_list_path: str, *, profile: dict | None = None) -> list[dict]:
    """
    Executor breakdown: same columns as streamwise but grouped by executor name.
    Cycles with multiple executors (comma-separated) are counted once per executor.
    Last row is the "Total" summary.
    """
    cycles = read_cycle_data(cycle_list_path, profile=profile)
    # Expand multi-executor cycles
    expanded: list[dict] = []
    for cyc in cycles:
        if cyc["executor"]:
            for name in [n.strip() for n in str(cyc["executor"]).split(",") if n.strip()]:
                expanded.append({**cyc, "executor": name})
        else:
            expanded.append({**cyc, "executor": "Unassigned"})
    return _workstream_summary(expanded, "executor")


def build_cycle_step_report(cycle_list_path: str, *, profile: dict | None = None) -> list[dict]:
    """
    Per-cycle progress table.
    Returns a list of dicts suitable for display as a dataframe.
    """
    cycles = read_cycle_data(cycle_list_path, profile=profile)
    rows = []
    for cyc in cycles:
        total    = cyc["total_steps"]    or 0
        complete = cyc["complete_steps"] or 0
        pct = f"{int(complete/total*100)}%" if total else "—"
        rows.append({
            "Cycle ID":      cyc["cycle_id"],
            "Name":          cyc["cycle_name"],
            "Area":          cyc["area"],
            "Executor":      cyc["executor"] or "—",
            "Status":        cyc["exec_status"],
            "Steps":         f"{complete}/{total} ({pct})",
            "Plan Start":    str(cyc["plan_start"])  if cyc["plan_start"]  else "—",
            "Plan End":      str(cyc["plan_end"])    if cyc["plan_end"]    else "—",
            "Actual Start":  str(cyc["actual_start"])if cyc["actual_start"]else "—",
            "Actual End":    str(cyc["actual_end"])  if cyc["actual_end"]  else "—",
            "Review End":    str(cyc["review_end"])  if cyc["review_end"]  else "—",
        })
    return rows


def build_merged_condition_report(script_paths: list[str], *,
                                   profile: dict | None = None) -> list[dict]:
    """
    Read all condition scripts and merge every active step into a flat list.
    Used to generate the "Condensed / Merged Condition File" output.

    Each row represents one test step and includes:
      cycle_id, step_no, executor, planned_date, actual_date, result,
      retest_date, retest_result, review_date, review_result
    """
    p = profile or get_active_profile()
    c = p["script_cols"]
    hdr = p["script_header_rows"]
    sheet_name = p["script_sheet_name"]
    all_steps: list[dict] = []

    def _fmt(val) -> str:
        if isinstance(val, datetime): return val.date().isoformat()
        if isinstance(val, date):     return val.isoformat()
        return str(val).strip() if val else ""

    for filepath in script_paths:
        cycle_id = extract_cycle_id_from_filename(Path(filepath).name, profile=p)
        if not cycle_id:
            continue
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception:
            continue
        if sheet_name not in wb.sheetnames:
            wb.close()
            continue
        ws = wb[sheet_name]
        last_step_no = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < hdr:
                continue
            if row and row[0] == "e":
                break
            if not row or all(v is None for v in row):
                continue
            if len(row) > c["excluded"] and row[c["excluded"]] == "X":
                continue

            step_no = row[1] if len(row) > 1 else None
            if step_no is not None:
                last_step_no = step_no

            has_data = (
                (len(row) > c["result"] and row[c["result"]] in ("OK", "NG", "-"))
                or (len(row) > c["actual_date"] and row[c["actual_date"]] is not None)
                or (len(row) > c["planned_date"] and row[c["planned_date"]] is not None)
            )
            if last_step_no is None and not has_data:
                continue

            def _get(col_key):
                idx = c.get(col_key)
                return row[idx] if idx is not None and idx < len(row) else None

            retest_result = _get("retest_result")
            first_result  = _get("result")
            result = (retest_result
                      if isinstance(retest_result, str) and retest_result.strip() not in ("", "-")
                      else first_result)

            all_steps.append({
                "Cycle ID":      cycle_id,
                "Step No":       last_step_no,
                "Executor":      _fmt(_get("exec_pic")),
                "Planned Date":  _fmt(_get("planned_date")),
                "Actual Date":   _fmt(_get("actual_date")),
                "Result":        _fmt(result) or "-",
                "Retest Date":   _fmt(_get("retest_date")),
                "Retest Result": _fmt(_get("retest_result")) or "-",
                "Review Date":   _fmt(_get("review_date")),
                "Review Result": _fmt(_get("review_result")) or "-",
            })
        wb.close()

    return all_steps


def get_executor_reminder_data(cycle_list_path: str, today: date, *,
                                profile: dict | None = None) -> dict[str, dict]:
    """
    Return per-executor data for the selective reminder UI.

    Structure:
      {
        "executor_name": {
          "exec_cycles":   [cycle dicts needing execution attention],
          "review_cycles": [cycle dicts needing review attention],
        }
      }

    Execution attention:
      - Status not yet started / in_progress / SIR
      - plan_end <= today (overdue or due today) OR plan_start in (today, tomorrow)

    Review attention:
      - Status = reviewing (in review)
      - review_plan_end <= today OR review_plan_end == tomorrow
    """
    tomorrow = today + timedelta(days=1)
    cycles = read_cycle_data(cycle_list_path, profile=profile)
    result: dict[str, dict] = {}

    for cyc in cycles:
        if cyc["exec_status"] == STATUS["cancelled"]:
            continue

        executor = cyc.get("executor")
        if not executor:
            continue

        for name in [n.strip() for n in str(executor).split(",") if n.strip()]:
            if name not in result:
                result[name] = {"exec_cycles": [], "review_cycles": []}

            # Execution attention: not yet complete and either overdue or starting soon
            is_exec_done = cyc["exec_status"] in (STATUS["complete"], STATUS["reviewing"])
            if not is_exec_done:
                needs_exec = (
                    (cyc["plan_end"] and cyc["plan_end"] <= today)        # overdue / due today
                    or (cyc["plan_start"] and cyc["plan_start"] in (today, tomorrow))  # starting
                )
                if needs_exec:
                    result[name]["exec_cycles"].append(cyc)

            # Review attention: in-review stage, review due or overdue
            if cyc["exec_status"] == STATUS["reviewing"]:
                needs_review = (
                    (cyc["review_plan_end"] and cyc["review_plan_end"] <= today)
                    or (cyc["review_plan_end"] and cyc["review_plan_end"] == tomorrow)
                )
                if needs_review:
                    result[name]["review_cycles"].append(cyc)

    # Remove executors with no pending cycles
    return {k: v for k, v in result.items() if v["exec_cycles"] or v["review_cycles"]}


def build_review_reminder(cycle: dict) -> str:
    """Build an HTML review reminder email for a reviewer."""
    cid  = cycle.get("cycle_id", "")
    name = cycle.get("cycle_name", "")
    rpe  = cycle.get("review_plan_end")
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:700px">
    <h2 style="color:#1565c0">🔍 Review Reminder — {cid}</h2>
    <p style="font-size:14px">
      Test cycle <b>{cid} — {name}</b> has completed execution and is awaiting review.
      {"Planned review deadline: <b>" + str(rpe) + "</b>.<br>" if rpe else ""}
      <br>
      Please review the condition script, verify the test evidence,
      and enter the review completion date once done.
    </p>
    <p style="font-size:13px;color:gray">
      This is an automated reminder from the SAP Test Management system.<br>
      Please do not reply to this email.
    </p>
    </div>"""


def build_remark_reminder(cycle: dict) -> str:
    """Build an HTML reminder to fill in remarks/notes in the condition script."""
    cid  = cycle.get("cycle_id", "")
    name = cycle.get("cycle_name", "")
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:700px">
    <h2 style="color:#6a1b9a">📝 Remark Filling Reminder — {cid}</h2>
    <p style="font-size:14px">
      Please make sure all remarks and notes are filled in for test cycle
      <b>{cid} — {name}</b> in the condition script.<br><br>
      Remarks help reviewers understand the test context and any issues encountered.
    </p>
    <p style="font-size:13px;color:gray">
      This is an automated reminder from the SAP Test Management system.<br>
      Please do not reply to this email.
    </p>
    </div>"""


# ─────────────────────────────────────────────────────────────
# EMAIL VIA LOCAL OUTLOOK (win32com — no API needed)
# ─────────────────────────────────────────────────────────────
def send_email_outlook(to: str, subject: str, body_html: str):
    """
    Send an email using the locally installed Outlook app via win32com.
    No API, no auth — uses your already-logged-in Outlook session.
    """
    try:
        import win32com.client
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = olMailItem
        mail.To      = to
        mail.Subject = subject
        mail.HTMLBody = body_html
        mail.Send()
        log.info(f"Email sent → {to}")
    except ImportError:
        log.error("win32com not installed. Run: pip install pywin32")
    except Exception as e:
        log.error(f"Failed to send email to {to}: {e}")


def resolve_email(executor_name: str) -> str | None:
    """
    Look up executor email from the name→email mapping in CONFIG.
    In test_mode, always returns the manager email so nothing goes to real executors.
    """
    if not executor_name:
        return None

    # TEST MODE: redirect everything to manager
    if CONFIG.get("test_mode"):
        return CONFIG["manager_email"]

    mapping = CONFIG["executor_emails"]
    # Try exact match first
    if executor_name in mapping:
        return mapping[executor_name]
    # Try partial match (e.g. "Tanaka K." matches "Tanaka")
    for name, email in mapping.items():
        if name.lower() in executor_name.lower():
            return email
    return None


def send_teams_channel(subject: str, body_html: str):
    """
    Post a message to the configured Teams channel via its email address.

    Every Teams channel has a built-in email address (channel → ••• → Get email
    address). Sending an email to that address creates a card in the channel —
    no webhook, no Azure AD app registration, no IT approval needed.

    The same win32com Outlook send is reused, so no extra dependencies are added.
    Messages appear as email cards in the Teams channel feed.
    """
    ch_email = CONFIG.get("teams_channel_email", "").strip()
    if not ch_email:
        log.warning("Teams notification skipped — teams_channel_email is not configured.")
        return
    send_email_outlook(ch_email, subject, body_html)
    log.info(f"Teams channel notified → {ch_email}")


def notify(to_email: str | None, subject: str, body_html: str, *, post_to_teams: bool = False):
    """
    Dispatch a notification to all configured channels.

    Parameters
    ----------
    to_email     : recipient for the "email" channel; pass None to skip email
    subject      : email / Teams card subject line
    body_html    : HTML body (rendered in Outlook and as Teams email card)
    post_to_teams: if True AND "teams" is in notify_channels, also posts to
                   the Teams channel. Use True for manager-level broadcasts;
                   False for per-executor messages that shouldn't flood the channel.

    Channel selection is controlled by CONFIG["notify_channels"].
    A future UI layer will expose this choice per run without changing CONFIG.
    """
    channels = CONFIG.get("notify_channels", ["email"])

    if CONFIG.get("dry_run"):
        log.info(f"[DRY RUN] Skipping notification: to={to_email!r}, subject={subject!r}, teams={post_to_teams}")
        return

    if "email" in channels and to_email:
        send_email_outlook(to_email, subject, body_html)

    if "teams" in channels and post_to_teams:
        send_teams_channel(subject, body_html)


# ─────────────────────────────────────────────────────────────
# REPORT BUILDER
# ─────────────────────────────────────────────────────────────
def build_manager_report(changes: list[dict], reminders: list[dict], today: date,
                          missing_scripts: list[dict] | None = None) -> str:
    """Build an HTML daily summary email for the manager."""
    missing_scripts = missing_scripts or []
    completed       = [c for c in changes  if c["new_status"] == STATUS["complete"]]
    in_review       = [c for c in changes  if c["new_status"] == STATUS["reviewing"]]
    started         = [c for c in changes  if c["new_status"] == STATUS["in_progress"]
                                           and c["old_status"] == STATUS["not_started"]]
    overdue         = [r for r in reminders if r["type"] == "overdue"]
    starting_today  = [r for r in reminders if r["type"] == "starting_today"]
    st_tomorrow     = [r for r in reminders if r["type"] == "starting_tomorrow"]
    due_tomorrow    = [r for r in reminders if r["type"] == "due_tomorrow"]

    def table(rows, cols):
        html = '<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-size:13px">'
        html += "<tr>" + "".join(f'<th style="background:#f0f0f0">{c}</th>' for c in cols) + "</tr>"
        for r in rows:
            html += "<tr>" + "".join(f"<td>{v}</td>" for v in r) + "</tr>"
        html += "</table>"
        return html

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:900px">
    <h2 style="color:#4a4a8a">📊 SAP Test Cycle Daily Report — {today.strftime('%Y/%m/%d')}</h2>
    <p style="font-size:15px">
      <b>✅ Completed:</b> {len(completed)} &nbsp;|&nbsp;
      <b>🔍 In Review:</b> {len(in_review)} &nbsp;|&nbsp;
      <b>🚀 Started:</b> {len(started)} &nbsp;|&nbsp;
      <b>🔴 Overdue:</b> {len(overdue)} &nbsp;|&nbsp;
      <b>🟢 Starting Today:</b> {len(starting_today)} &nbsp;|&nbsp;
      <b>⚠️ No script:</b> {len(missing_scripts)}
    </p>
    """

    if completed:
        html += "<h3 style='color:green'>✅ Completed (Review Done)</h3>"
        html += table(
            [(c["cycle_id"], c["cycle_name"], c["area"],
              f"{c['completed_steps']}/{c['total_steps']}", c["executor"] or "-")
             for c in completed],
            ["Cycle ID", "Name", "Area", "Steps", "Executor"]
        )

    if in_review:
        html += "<h3 style='color:#0077cc'>🔍 Moved to In Review</h3>"
        html += table(
            [(c["cycle_id"], c["cycle_name"], c["area"],
              f"{c['completed_steps']}/{c['total_steps']}", c["executor"] or "-")
             for c in in_review],
            ["Cycle ID", "Name", "Area", "Steps", "Executor"]
        )

    if overdue:
        html += "<h3 style='color:red'>🔴 Overdue — Action Required</h3>"
        html += table(
            [(r["cycle_id"], r["cycle_name"], str(r["plan_end"] or "-"),
              r["executor"] or "-", r["status"])
             for r in overdue],
            ["Cycle ID", "Name", "Plan End", "Executor", "Status"]
        )

    if due_tomorrow:
        html += "<h3 style='color:orange'>⚡ Due Tomorrow</h3>"
        html += table(
            [(r["cycle_id"], r["cycle_name"], str(r["plan_end"] or "-"),
              r["executor"] or "-", r["status"])
             for r in due_tomorrow],
            ["Cycle ID", "Name", "Plan End", "Executor", "Status"]
        )

    if starting_today:
        html += "<h3 style='color:#2e7d32'>🟢 Starting Today</h3>"
        html += table(
            [(r["cycle_id"], r["cycle_name"], str(r["plan_start"] or "-"), r["executor"] or "-")
             for r in starting_today],
            ["Cycle ID", "Name", "Plan Start", "Executor"]
        )

    if st_tomorrow:
        html += "<h3 style='color:#1565c0'>📅 Starting Tomorrow</h3>"
        html += table(
            [(r["cycle_id"], r["cycle_name"], str(r["plan_start"] or "-"), r["executor"] or "-")
             for r in st_tomorrow],
            ["Cycle ID", "Name", "Plan Start", "Executor"]
        )

    if started:
        html += "<h3>🚀 Newly Started</h3>"
        html += table(
            [(c["cycle_id"], c["cycle_name"], c["area"],
              f"{c['completed_steps']}/{c['total_steps']}", c["executor"] or "-")
             for c in started],
            ["Cycle ID", "Name", "Area", "Steps", "Executor"]
        )

    if missing_scripts:
        html += "<h3 style='color:#b8860b'>⚠️ No Script File Found</h3>"
        html += table(
            [(m["cycle_id"], m["cycle_name"], m["area"] or "-",
              m["exec_status"] or "-", m["executor"] or "-")
             for m in missing_scripts],
            ["Cycle ID", "Name", "Area", "Current Status", "Executor"]
        )

    html += "<br><p style='color:gray;font-size:11px'>Generated automatically by SAP Test Automation</p></div>"
    return html


def build_executor_reminder(reminder: dict) -> str:
    """Build an HTML reminder email for an individual executor."""
    r = reminder
    if r["type"] == "overdue":
        heading = f"🔴 Overdue Test Cycle — {r['cycle_id']}"
        msg = f"""Your test cycle <b>{r['cycle_id']} — {r['cycle_name']}</b> was due by
        <b>{r['plan_end']}</b> and has not been completed.<br><br>
        Please update the condition script with your execution results (actual dates + OK/NG)
        as soon as possible. Current status: <b>{r['status']}</b>."""
    elif r["type"] == "starting_today":
        heading = f"🟢 Test Cycle Starts Today — {r['cycle_id']}"
        msg = f"""Your test cycle <b>{r['cycle_id']} — {r['cycle_name']}</b> is scheduled
        to start <b>today ({r['plan_start']})</b>.<br><br>
        Please make sure your test data and SAP environment are ready.
        Check the condition script and confirm all pre-conditions are met before starting."""
    elif r["type"] == "starting_tomorrow":
        heading = f"📅 Test Cycle Starts Tomorrow — {r['cycle_id']}"
        msg = f"""Your test cycle <b>{r['cycle_id']} — {r['cycle_name']}</b> is scheduled
        to start <b>tomorrow ({r['plan_start']})</b>.<br><br>
        Please prepare your test environment and review the condition script today
        so you can start on time tomorrow."""
    elif r["type"] == "due_tomorrow":
        heading = f"⚡ Test Cycle Due Tomorrow — {r['cycle_id']}"
        msg = f"""Your test cycle <b>{r['cycle_id']} — {r['cycle_name']}</b> is due for
        completion by <b>tomorrow ({r['plan_end']})</b>.<br><br>
        Please ensure all steps are executed and results (actual dates + OK/NG) are entered
        in the condition script before end of day. Current status: <b>{r['status']}</b>."""
    else:
        heading = f"🔄 Please Update Test Results — {r['cycle_id']}"
        msg = f"""Test cycle <b>{r['cycle_id']} — {r['cycle_name']}</b> is in progress
        but results have not been fully entered.<br><br>
        Please open the condition script and enter the actual execution date and
        OK/NG result for each completed step."""

    return f"""
    <div style="font-family:Arial,sans-serif;max-width:700px">
    <h2 style="color:#4a4a8a">{heading}</h2>
    <p style="font-size:14px">{msg}</p>
    <p style="font-size:13px;color:gray">
      This is an automated reminder from the SAP Test Management system.<br>
      Please do not reply to this email.
    </p>
    </div>
    """


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def run(time_slot: str = "morning"):
    """
    Main orchestration function.

    Parameters
    ----------
    time_slot : str
        Controls which reminder types are sent:
          "morning"  — overdue + starting today            (for 9:00 run)
          "midday"   — same as morning                     (for 12:00 run)
          "evening"  — morning + starting tomorrow + due tomorrow  (for 17:00 run)
    """
    setup_logging()
    today = date.today()
    log.info(f"{'='*60}")
    log.info(f"SAP Test Automation — {today}")
    log.info(f"{'='*60}")

    # Resolve the active profile once; pass it to every function so they all
    # use the same consistent configuration for this run.
    profile = get_active_profile()
    log.info(f"Active profile: {profile.get('display_name', CONFIG.get('active_profile'))}")
    if not profile.get("configured", True):
        log.warning("Profile is not fully configured — column indices may be incorrect. "
                    "Verify via the Streamlit profile editor before running live.")

    folder = CONFIG["folder"]
    if not os.path.exists(folder):
        log.error(f"Folder not found: {folder}")
        log.error("Check the 'folder' path in CONFIG and make sure OneDrive is synced.")
        return

    # 1. Find files
    cycle_list_path, script_paths = find_files(folder, profile=profile)

    if not cycle_list_path:
        log.error("Could not find the Test Cycle List file. Check 'cycle_list_pattern' in the profile.")
        return

    if not script_paths:
        log.error("No condition script files found. Check 'condition_script_pattern' in the profile.")
        return

    # 2. Analyze all condition scripts
    log.info(f"Analyzing {len(script_paths)} condition scripts...")
    script_results = {}
    for path in script_paths:
        result = analyze_script(path, profile=profile)
        if result:
            script_results[result["cycle_id"]] = result
            status_str = "✅ DONE" if result["all_ok"] else f"{result['completed_steps']}/{result['total_steps']} steps"
            log.info(f"  {result['cycle_id']}: {status_str} {'⚠️ HAS NG' if result['has_ng'] else ''}")

    log.info(f"Successfully analyzed: {len(script_results)} scripts")

    # 3. Update cycle list
    log.info("Updating Test Cycle List...")
    changes, missing_scripts = update_cycle_list(cycle_list_path, script_results, profile=profile)
    if missing_scripts:
        log.warning(f"{len(missing_scripts)} active cycle(s) have no matching script:")
        for m in missing_scripts:
            log.warning(f"  No script: {m['cycle_id']}  ({m['cycle_name']})")
    # OneDrive will auto-sync the updated file to SharePoint

    # 4. Get reminders (criteria depend on the time slot)
    log.info(f"Time slot: {time_slot}")
    reminders = get_reminders(cycle_list_path, today, profile=profile, time_slot=time_slot)
    log.info(f"Reminders: {len(reminders)} cycles need attention")

    # 5. Send individual executor reminders
    sent_to = set()
    for reminder in reminders:
        executor_name = reminder.get("executor")
        if not executor_name:
            continue

        # Handle multiple executors in one field (e.g. "Tanaka, Suzuki")
        for name in [n.strip() for n in str(executor_name).split(",")]:
            email = resolve_email(name)
            if not email:
                log.warning(f"No email mapping for executor: '{name}' — add to CONFIG['executor_emails']")
                continue

            # Avoid duplicate emails to same person
            reminder_key = f"{email}_{reminder['cycle_id']}_{reminder['type']}"
            if reminder_key in sent_to:
                continue
            sent_to.add(reminder_key)

            subject = f"[SAP Test{'  TEST MODE' if CONFIG.get('test_mode') else ''}] {reminder['type'].upper()} — {reminder['cycle_id']} (executor: {name})"
            body    = build_executor_reminder(reminder)
            # post_to_teams=False: executor reminders go to email only to avoid
            # flooding the shared Teams channel with per-person messages.
            notify(email, subject, body, post_to_teams=False)

    # 6. Send daily summary to manager (email + Teams channel)
    report_html = build_manager_report(changes, reminders, today, missing_scripts)
    notify(
        CONFIG["manager_email"],
        f"[SAP Test] Daily Report {today.strftime('%Y/%m/%d')} — "
        f"{len([c for c in changes if c['new_status'] == STATUS['complete']])} completed, "
        f"{len([r for r in reminders if r['type'] == 'overdue'])} overdue",
        report_html,
        post_to_teams=True,   # manager summary always goes to the Teams channel
    )

    log.info("=== Run complete ===")
    log.info(f"Changes: {len(changes)} | Reminders sent: {len(sent_to)} | Report: sent to {CONFIG['manager_email']}")


if __name__ == "__main__":
    import sys
    # Optionally pass a time slot as the first argument:
    #   python AT_Sentinel.py morning   → overdue + starting today
    #   python AT_Sentinel.py midday    → same as morning
    #   python AT_Sentinel.py evening   → morning + starting tomorrow + due tomorrow
    slot = sys.argv[1] if len(sys.argv) > 1 else "morning"
    run(time_slot=slot)
